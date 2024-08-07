import openpyxl as xl  # Import the openpyxl library for working with Excel files
from openpyxl.chart import BarChart, Reference  # Import classes for creating charts

filename = 'transactions.xlsx'  # Define the filename of the Excel file to be processed

def process_workbook(filename):
    wb = xl.load_workbook(filename)  # Load the workbook from the specified filename
    sheet = wb['Sheet1']  # Select the sheet named 'Sheet1'

    # Iterate over rows starting from the second row to the last row
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # Access the cell in the third column of the current row
        
        # Calculate the corrected price by applying a 10% discount
        corrected_price = cell.value * 0.9
        
        # Access the cell in the fourth column of the current row
        corrected_price_cell = sheet.cell(row, 4)
        
        # Set the corrected price in the fourth column cell
        corrected_price_cell.value = corrected_price

    # Create a reference to the range of cells to include in the chart (fourth column, from row 2 to the last row)
    values = Reference(sheet,
                        min_row=2, 
                        max_row=sheet.max_row,
                        min_col=4,
                        max_col=4)

    # Create a BarChart object
    chart = BarChart()
    
    # Add the data reference to the chart
    chart.add_data(values)
    
    # Add the chart to the sheet at cell 'E2'
    sheet.add_chart(chart, 'E2')

    # Save the modified workbook with a new filename to test the changes
    wb.save('transactions2.xlsx')  # Use this to test if the program is working well before overwriting the original Excel file
    # wb.save(filename)  # Uncomment this to save changes to the original file

# Call the function to process the workbook with the specified filename
process_workbook(filename)
