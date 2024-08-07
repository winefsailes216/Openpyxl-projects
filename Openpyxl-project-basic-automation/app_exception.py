import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Define the filename for the Excel file to be processed
filename = 'transactions.xlsx'

def process_workbook(filename):
    try:
        # Load the workbook and select the sheet
        wb = xl.load_workbook(filename)
        sheet = wb['Sheet1']
    except FileNotFoundError:
        print(f"Error: The file '{filename}' was not found.")
        return
    except KeyError:
        print("Error: The sheet 'Sheet1' does not exist in the workbook.")
        return

    try:
        # Iterate over rows starting from the second row
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            
            # Check if the cell value is numeric
            if isinstance(cell.value, (int, float)):
                # Calculate the corrected price by applying a 10% discount
                corrected_price = cell.value * 0.9
                
                # Write the corrected price into the fourth column of the same row
                corrected_price_cell = sheet.cell(row, 4)
                corrected_price_cell.value = corrected_price
            else:
                print(f"Warning: Non-numeric value encountered in row {row}. Skipping this row.")

        # Create a reference to the range of cells to include in the chart
        values = Reference(sheet,
                            min_row=2, 
                            max_row=sheet.max_row,
                            min_col=4,
                            max_col=4)

        # Create a BarChart object and add data to it
        chart = BarChart()
        chart.add_data(values)
        
        # Add the chart to the sheet at the specified location
        sheet.add_chart(chart, 'E2')

        # Save the workbook with a new filename to avoid overwriting the original
        wb.save('transactions3.xlsx')  # Use this to test if the program is working well before overwriting the original Excel file
        # wb.save(filename)  # Uncomment this to overwrite the original file
    
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Call the function to process the workbook
process_workbook(filename)
